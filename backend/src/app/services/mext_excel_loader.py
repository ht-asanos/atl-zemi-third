"""MEXT 食品成分 Excel ファイルパーサー

文科省「日本食品標準成分表（八訂）」の Excel ファイルから
MextFood オブジェクトを生成する。

使用方法:
    foods, stats = load_foods_from_excel("path/to/012.xlsx")
"""

from __future__ import annotations

import logging
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
from app.models.food import MextFood

logger = logging.getLogger(__name__)

# ヘッダーキーワード → 内部フィールド名
HEADER_KEYWORDS: dict[str, str] = {
    "食品名": "name",
    "エネルギー": "_energy_start",
    "たんぱく質": "_protein_area",
    "脂質": "_fat_area",
    "炭水化物": "_carbs_area",
    "食物繊維総量": "fiber",
    "ナトリウム": "sodium",
    "カルシウム": "calcium",
    "鉄": "iron",
}

# フォールバック: 012.xlsx の固定列インデックス (1-based)
FALLBACK_COLUMNS: dict[str, int] = {
    "category_code": 1,
    "mext_food_id": 2,
    "name": 4,
    "kcal": 7,
    "protein": 10,
    "fat": 13,
    "carbs": 21,
    "fiber": 19,
    "sodium": 24,
    "calcium": 26,
    "iron": 29,
}

# 成分識別子による列検出 (Row 5 前後に記載)
COMPONENT_ID_MAP: dict[str, str] = {
    "ENERC_KCAL": "kcal",
    "PROT-": "protein",
    "FAT-": "fat",
    "CHOCDF-": "carbs",
    "FIB-": "fiber",
    "NA": "sodium",
    "CA": "calcium",
    "FE": "iron",
}

# 食品群コード → カテゴリ名
CATEGORY_NAMES: dict[str, str] = {
    "01": "穀類",
    "02": "いも及びでん粉類",
    "03": "砂糖及び甘味類",
    "04": "豆類",
    "05": "種実類",
    "06": "野菜類",
    "07": "果実類",
    "08": "きのこ類",
    "09": "藻類",
    "10": "魚介類",
    "11": "肉類",
    "12": "卵類",
    "13": "乳類",
    "14": "油脂類",
    "15": "菓子類",
    "16": "し好飲料類",
    "17": "調味料及び香辛料類",
    "18": "調理加工食品類",
    "19": "その他",
}

# ヘッダー走査範囲
HEADER_SCAN_MIN_ROW = 1
HEADER_SCAN_MAX_ROW = 12
DATA_START_ROW = 13


def _normalize_header(val: Any) -> str:
    """セル値を正規化して比較用文字列にする。"""
    if val is None:
        return ""
    s = str(val)
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("\n", "").replace("\r", "").replace(" ", "").replace("\u3000", "")
    return s


def _detect_columns(ws: Any) -> dict[str, int]:
    """ヘッダー行を走査してフィールド→列インデックスのマッピングを構築する。"""
    detected: dict[str, int] = {
        "category_code": FALLBACK_COLUMNS["category_code"],
        "mext_food_id": FALLBACK_COLUMNS["mext_food_id"],
    }
    fallback_used: list[str] = []

    # 成分識別子行から列を検出
    for row in ws.iter_rows(min_row=HEADER_SCAN_MIN_ROW, max_row=HEADER_SCAN_MAX_ROW, values_only=False):
        for cell in row:
            if not hasattr(cell, "column"):
                continue
            val = _normalize_header(cell.value)
            if not val:
                continue
            if val in COMPONENT_ID_MAP:
                field = COMPONENT_ID_MAP[val]
                if field not in detected:
                    detected[field] = cell.column

    # キーワードベースの検出 (成分識別子で見つからなかった列)
    for row in ws.iter_rows(min_row=HEADER_SCAN_MIN_ROW, max_row=HEADER_SCAN_MAX_ROW, values_only=False):
        for cell in row:
            if not hasattr(cell, "column"):
                continue
            val = _normalize_header(cell.value)
            if not val:
                continue
            for keyword, field_name in HEADER_KEYWORDS.items():
                if keyword in val and field_name not in detected and not field_name.startswith("_"):
                    detected[field_name] = cell.column

    # name 列の検出: 「食品名」キーワードから
    if "name" not in detected:
        for row in ws.iter_rows(min_row=HEADER_SCAN_MIN_ROW, max_row=HEADER_SCAN_MAX_ROW, values_only=False):
            for cell in row:
                if not hasattr(cell, "column"):
                    continue
                val = _normalize_header(cell.value)
                if "食品名" in val:
                    detected["name"] = cell.column
                    break
            if "name" in detected:
                break

    # フォールバック適用
    for field, col_idx in FALLBACK_COLUMNS.items():
        if field not in detected:
            detected[field] = col_idx
            fallback_used.append(field)

    if fallback_used:
        logger.warning("Fallback columns used: %s", fallback_used)
    logger.info("Column mapping: %s", detected)

    return detected


def _parse_float(val: Any) -> tuple[float, str]:
    """(数値, フラグ) を返す。フラグ: 'value' | 'trace' | 'missing'"""
    if val is None:
        return 0.0, "missing"
    s = str(val).strip()
    s = s.replace("(", "").replace(")", "").replace("（", "").replace("）", "")
    s = s.replace(",", "")
    if s in ("", "-", "—", "*", "…"):
        return 0.0, "missing"
    if s in ("Tr", "tr", "TR"):
        return 0.0, "trace"
    try:
        return float(s), "value"
    except ValueError:
        return 0.0, "missing"


def _cell_value(row_cells: dict[int, Any], col: int) -> Any:
    """行のセル辞書から値を取得する。"""
    return row_cells.get(col)


def load_foods_from_excel(path: str | Path) -> tuple[list[MextFood], dict[str, Any]]:
    """Excel から MextFood リストを生成する。

    Returns:
        (foods, stats) タプル。stats にはパース統計を含む。
    """
    path = Path(path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    col_map = _detect_columns(ws)
    fallback_cols = [
        field
        for field, col_idx in FALLBACK_COLUMNS.items()
        if col_map.get(field) == col_idx and field not in ("category_code", "mext_food_id")
    ]

    foods: list[MextFood] = []
    stats: dict[str, Any] = {
        "total_rows": 0,
        "skipped": 0,
        "zero_kcal": 0,
        "fallback_cols": fallback_cols,
        "file": path.name,
    }

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=False):
        row_cells: dict[int, Any] = {cell.column: cell.value for cell in row if hasattr(cell, "column")}

        category_code = str(_cell_value(row_cells, col_map["category_code"]) or "").strip()
        mext_food_id = str(_cell_value(row_cells, col_map["mext_food_id"]) or "").strip()
        name = str(_cell_value(row_cells, col_map["name"]) or "").strip()

        if not mext_food_id or not name:
            stats["skipped"] += 1
            continue

        stats["total_rows"] += 1

        kcal_val, kcal_flag = _parse_float(_cell_value(row_cells, col_map["kcal"]))
        protein_val, protein_flag = _parse_float(_cell_value(row_cells, col_map["protein"]))
        fat_val, fat_flag = _parse_float(_cell_value(row_cells, col_map["fat"]))
        carbs_val, carbs_flag = _parse_float(_cell_value(row_cells, col_map["carbs"]))
        fiber_val, fiber_flag = _parse_float(_cell_value(row_cells, col_map["fiber"]))
        sodium_val, sodium_flag = _parse_float(_cell_value(row_cells, col_map["sodium"]))
        calcium_val, calcium_flag = _parse_float(_cell_value(row_cells, col_map["calcium"]))
        iron_val, iron_flag = _parse_float(_cell_value(row_cells, col_map["iron"]))

        trace_fields = []
        missing_fields = []
        for field_name, flag in [
            ("kcal_per_100g", kcal_flag),
            ("protein_g_per_100g", protein_flag),
            ("fat_g_per_100g", fat_flag),
            ("carbs_g_per_100g", carbs_flag),
            ("fiber_g_per_100g", fiber_flag),
            ("sodium_mg_per_100g", sodium_flag),
            ("calcium_mg_per_100g", calcium_flag),
            ("iron_mg_per_100g", iron_flag),
        ]:
            if flag == "trace":
                trace_fields.append(field_name)
            elif flag == "missing":
                missing_fields.append(field_name)

        if kcal_val == 0 and kcal_flag != "trace":
            stats["zero_kcal"] += 1

        category_name = CATEGORY_NAMES.get(category_code, category_code)

        raw_data = {
            "source": "excel",
            "file": path.name,
            "trace_fields": trace_fields,
            "missing_fields": missing_fields,
        }

        food = MextFood(
            mext_food_id=mext_food_id,
            name=name,
            category_code=category_code,
            category_name=category_name,
            kcal_per_100g=kcal_val,
            protein_g_per_100g=protein_val,
            fat_g_per_100g=fat_val,
            carbs_g_per_100g=carbs_val,
            fiber_g_per_100g=fiber_val,
            sodium_mg_per_100g=sodium_val,
            calcium_mg_per_100g=calcium_val,
            iron_mg_per_100g=iron_val,
            raw_data=raw_data,
        )
        foods.append(food)

    wb.close()
    return foods, stats
